import xlrd
import openpyxl
from logging import getLogger
import datetime
from datetime import datetime


class ReadExcel(object):
    def __init__(self, file_path):
        self.file = file_path
        self.work_book = xlrd.open_workbook(file_path)
        self.sheets = self.work_book.sheets()

    def read_sheet_name(self) -> list:
        """
        读取sheet名称，返回一个列表
        :return: [sheet1, sheet2...]
        """
        sheet_name_list = [i.name for i in self.sheets]
        return sheet_name_list

    def read_table_head(self, sheet_index=None) -> list:
        """
        按sheet_index读取第几个sheet的表头，默认从第一个往后读到有表头的数据
        :return: [query, search_location...]
        """
        log = getLogger("read_table_head")
        if sheet_index == 0 or sheet_index:
            try:
                sheet = self.sheets[sheet_index]
                first_rows = [i.value for i in sheet.row(0)]
                return first_rows
            except Exception as e:
                log.error("read_table_head Exception : {}".format(e))
                return []
        for sheet in self.sheets:
            rows = sheet.nrows
            if rows > 0:
                first_rows = sheet.row(0)
                return [i.value for i in first_rows]
        log.error("文件:{}, 未检测到表头".format(self.file))
        return []

    @staticmethod
    def _read_int(text):
        """
        针对整数，读出来会变成.0（1 -> 1.0）的情况做处理
        :type text: str
        :return:
        """
        if type(text) is float and '.0' == str(text)[-2:]:
            text = str(text).replace('.0', '')
        if len(text) > 2 and text[-2:] == '.0':
            text = text.replace('.0', '')
        if text.isdigit():
            return int(text)
        return text

    def _get_head_index(self, sheet_index, heads, heads_index) -> list:
        # sheet = self.sheets[sheet_index]
        if heads_index:
            if len(heads_index) == len(heads):
                return heads_index
            else:
                raise Exception("len(heads_index) 必须等于 len(heads)")
        table_head = self.read_table_head(sheet_index)
        result = []
        for head in heads:
            if head in table_head:
                result.append(table_head.index(head))
            else:
                result.append(-1)
        return result

    def read_sheet_data(self, heads, heads_index=None, sheet_index=0) -> list:
        """
        按给的head读取指定的sheet, 默认第0个
        :param heads: 表头
        :param heads_index: 表头所在的索引
        :param sheet_index: 读取第几个sheet
        :return:
        """
        log = getLogger("read_sheet")
        result = []
        try:
            sheet = self.sheets[sheet_index]
            rows = sheet.nrows
            if rows == 0:
                return []
            heads_index = self._get_head_index(sheet_index, heads, heads_index)
            for row in range(1, rows):
                row_dict = dict()
                row_data = sheet.row(row)
                for head, head_index in zip(heads, heads_index):
                    row_dict[head] = self._read_int(row_data[head_index].value)
                result.append(row_dict)
            return result
        except Exception as e:
            log.error("read_sheet_data Exception sheet_index = {}, error_info:{}".format(sheet_index, e))
            return []

    def read_sheets_data(self, heads, heads_index=None) -> dict:
        """
        按每个sheet读取数据
        :type heads: list
        :type heads_index: list
        :return:{sheet1: [row1, row2...], sheet2: [row1, row2]}
        """
        sheet_index = 0
        log = getLogger("read_sheets")
        result = {}
        try:
            for sheet in self.sheets:
                sheet_name = sheet.name
                rows = sheet.nrows
                if rows == 0:
                    continue
                result[sheet_name] = self.read_sheet_data(heads, heads_index, sheet_index)
                sheet_index += 1
            return result
        except Exception as e:
            log.error("read_sheets Exception : {}".format(e))
            return result

    def statistics_for_sheets(self) -> dict:
        """
        统计每个sheet名有多少条数据
        :return: {sheet1: 123, sheet2: 456}
        """
        result = {}
        log = getLogger("statistics_for_sheets")
        for sheet, sheet_name in zip(self.sheets, self.read_sheet_name()):
            rows = sheet.nrows
            if rows > 1:
                result[sheet_name] = rows - 1
            else:
                log.info("statistics_for_sheets info \"{}\"的行数为{}，不做统计".format(sheet_name, rows))
        return result

    def statistics_sheet_for_clo(self, col_name, sheet_index=0) -> dict:
        """
        按列明统计指定sheet，默认第0个
        :return: {clo_value1: 123, col_value2: 456...}
        """
        # log = getLogger("statistics_sheet_for_clo")
        result = {}
        try:
            heads = self.read_table_head(sheet_index)
            if col_name not in heads:
                raise Exception("未找到表头:{}, sheet_index:{}".format(col_name, sheet_index))
            sheet_data = self.read_sheet_data(heads, [], sheet_index)
            for row_dict in sheet_data:
                if row_dict.get(col_name, "") not in result:
                    result[row_dict[col_name]] = 1
                else:
                    result[row_dict[col_name]] += 1
            return result
        except Exception as e:
            raise Exception("statistics_sheet_for_clo Exception :{}".format(e))

    def statistics_sheets_for_col(self, col_name) -> dict:
        """
        按sheet和列明统计数据
        :param col_name: 要按哪一列统计
        :return: {sheet1: {col1: 123, clo2: 456...}, sheet2:{}...}
        """
        log = getLogger("statistics_sheets_for_col")
        result = {}
        sheet_index = 0
        for sheet, sheet_name in zip(self.sheets, self.read_sheet_name()):
            try:
                if sheet.nrows > 1:
                    sheet_data = self.statistics_sheet_for_clo(col_name, sheet_index)
                    result[sheet_name] = sheet_data
            except Exception as e:
                log.error("statistics_sheets_for_col Exception: {}".format(e))
            finally:
                sheet_index += 1
        return result


class WriteExcel(object):
    def __init__(self, write_result):
        self.log = getLogger("WriteExcel")
        if not write_result or write_result == [{}]:
            self.log.error("WriteExcel __init__ Exception write_result 没有可写入的数据")
            return
        self.work_book = openpyxl.Workbook()
        self.write_result = write_result
        now = datetime.now()
        now_date = now.strftime("%Y-%m-%d")
        self.save_name = '{}.xlsx'.format(now_date)
        self.work_book.remove(self.work_book['Sheet'])

    @staticmethod
    def _deal_with_num(_text):
        """
        针对整数，读出来会变成.0（1 -> 1.0）的情况做处理
        :type _text: str
        :return:
        """
        log = getLogger("_deal_with_num")
        try:

            text = _text
            if type(text) is int:
                return text
            if type(text) is float and '.0' == str(text)[-2:]:
                text = str(text).replace('.0', '')
            if type(text) is str and len(text) > 2 and text[-2:] == '.0':
                text = text.replace('.0', '')
            if type(text) is str and text.isdigit():
                return int(text)
            return text
        except Exception as e:
            log.error("_deal_with_num Exception :{}".format(e))
            return _text
        # return text

    def _get_save_name(self, name=None) -> str:
        if not name:
            return self.save_name
        elif '.xlsx' in name:
            return name
        else:
            return name + '.xlsx'

    def _get_heads(self, heads=None) -> list:
        """
        获取写入数据的表头
        :param heads: 表头
        :return:
        """
        if heads:
            return heads
        if type(self.write_result) is list:
            return [i for i in self.write_result[0].keys()]
        elif type(self.write_result) is dict:
            result = []
            for sheet, sheet_result in self.write_result.items():
                heads = []
                if isinstance(sheet_result, list):
                    heads = [i for i in sheet_result[0].keys()]
                elif isinstance(sheet_result, dict):
                    heads = [i for i in sheet_result.keys()]
                for head in heads:
                    if head not in result:
                        result.append(head)
            return result
        else:
            raise Exception("write_result only support list or dict")

    def write_sheet(self, write_result=None, heads=None, sheet_name=None, save_name=None, save_status=1):
        """
        写入单个sheet,写入内容必须是列表
        :param write_result:
        :param save_status: 是否保存文件，1：保存；0：不保存
        :param heads: 表头
        :type heads: list
        :param sheet_name: sheet名称
        :type sheet_name: str
        :param save_name: 最后保存的文件名
        :return:
        """
        log = getLogger("write_sheet")
        if not write_result:
            write_result = self.write_result
        save_name = self._get_save_name(save_name)
        if not isinstance(write_result, list):
            raise Exception("write_sheet Exception type of write_result is not list")
        sheet_name = sheet_name if sheet_name else 'Sheet'
        sheet = self.work_book.create_sheet(title=sheet_name)
        write_head = self._get_heads(heads)
        row = 1
        for head, head_index in zip(write_head, range(1, len(write_head) + 1)):
            try:
                sheet.cell(row=1, column=head_index, value=self._deal_with_num(head))
            except Exception as e:
                sheet.cell(row=1, column=head_index, value='')
                log.error("表头写入失败,{}, 请自行填写, row={}, column={}, value=\"{}\"".format(e, 1, head_index, head))
        for row_dict in write_result:
            row += 1
            col_num = 0
            for head in write_head:
                col_num += 1
                write_text = row_dict[head] if head in row_dict else ''
                try:
                    sheet.cell(row=row, column=col_num, value=self._deal_with_num(write_text))
                except Exception as e:
                    log.error("数据写入失败,{}, 请自行填写, row={}, column={}, value=\"{}\"".format(e, row, col_num, write_text))
        if save_status == 1:
            self.work_book.save(save_name)

    def write_sheets(self, heads=None, save_name=None):
        """
        写入多个sheet，write_result必须是dict
        :param heads: 表头
        :param save_name: 保存文件名
        :return:
        """
        heads = self._get_heads(heads)
        save_name = self._get_save_name(save_name)
        if not isinstance(self.write_result, dict):
            raise Exception("write_sheets Exception type of write_result is not dict")
        for sheet, sheet_result in self.write_result.items():
            self.write_sheet(sheet_result, heads, sheet, "", save_status=0)
        self.work_book.save(save_name)

    def write_statistics(self):
        """
        以统计的格式写入表格
        :return:
        """
        log = getLogger("write_statistics")
        if not isinstance(self.write_result, dict):
            raise Exception("write_statistics Exception only support dict")
        sheet = self.work_book.create_sheet(title='统计')
        heads = self._get_heads()
        # 写入表头
        for head, col_num in zip(heads, range(2, len(heads) + 2)):
            try:
                sheet.cell(row=1, column=col_num, value=head)
            except Exception as e:
                log.error("表头写入失败,{}, 请自行填写, row={}, column={}, value=\"{}\"".format(e, 1, col_num, head))
        row = 1
        # 写入数据
        for sheet_name in self.write_result:
            row += 1
            if not isinstance(self.write_result[sheet_name], dict):
                log.error(
                    'write_statistics Exception {}数据类型是{}，已跳过'.format(sheet_name, type(self.write_result[sheet_name]))
                )
                continue
            try:
                sheet.cell(row=row, column=1, value=sheet_name)
            except Exception as e:
                log.error("数据写入失败,{}, 请自行填写, row={}, column={}, value=\"{}\"".format(e, row, 1, sheet_name))
            col_num = 1
            for head in heads:
                col_num += 1
                try:
                    write_text = self.write_result[sheet_name][head] if head in self.write_result[sheet_name] else ''
                    sheet.cell(row=row, column=col_num, value=self._deal_with_num(write_text))
                except Exception as e:
                    log.error("数据写入失败,{}, 请自行填写, row={}, column={}, value=\"{}\"".format(e, row, 1, sheet_name))
        self.work_book.save(self.save_name)

    def __del__(self):
        self.work_book.close()


if __name__ == '__main__':
    file = r'C:\Users\Administrator\Desktop\test.xlsx'
    a = ReadExcel(file)
    # r_heads = ['query', 'pattern', 'label']
    # print(a.read_sheets_data(r_heads, [0, 1, 2]))
    print(a.statistics_sheets_for_col("pattern"))
    a = {'Sheet1': {1: 1, 2: 2}, 'Sheet2': {1: 1, 2: 1, 3: 1}}
    # a = [{"query": '南京', 'pattern': 2, 'label': '3'}, {'query': '北京', 'pattern': 3, 'label': 4}]
    b = WriteExcel(a)
    print(b.write_statistics())
    # print(b.write_sheets())
